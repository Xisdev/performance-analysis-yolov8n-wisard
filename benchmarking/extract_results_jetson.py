"""
Ekstrak data benchmark Jetson Nano: Latency, FPS, Resource (Average & Max), Power.
Warmup frames diexclude. Lalu validasi terhadap hasil manual Excel user.
"""
import os
import csv
import json

DATA_DIR = r"D:\RISET\drone-wisard\test_for_device\jetson_nano\data_hasil_jetson_baru\MENTAH_YANG_SUDAH-DIDATA"
OUTPUT_DIR = r"D:\RISET\drone-wisard\test_for_device\jetson_nano\data_hasil_jetson_baru\hasil"
EXCEL_DIR = r"D:\RISET\drone-wisard\test_for_device\jetson_nano\data_hasil_jetson_baru\FIXED-baru"
os.makedirs(OUTPUT_DIR, exist_ok=True)

MODELS = ['yolov8n', 'yolov8s', 'yolov8m', 'yolov8l', 'yolov8x']

def read_csv(path):
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return list(reader)

# ============================================================
# 1. LATENCY & FPS
# ============================================================
latency_results = {}

for model in MODELS:
    csv_file = os.path.join(DATA_DIR, f"benchmark_fps_latency_best_{model}_fp16_jetson_nano.csv")
    if not os.path.exists(csv_file):
        print(f"[SKIP] {csv_file} not found")
        continue

    rows = read_csv(csv_file)
    data = [r for r in rows if r.get('is_warmup', 'true') == 'false'
            and r.get('preprocess_ms', 'ERROR') not in ('ERROR', 'OOM')]

    if not data:
        print(f"[WARN] No valid data for {model}")
        continue

    preprocess = [float(r['preprocess_ms']) for r in data]
    inference = [float(r['inference_ms']) for r in data]
    postprocess = [float(r['postprocess_ms']) for r in data]
    total = [float(r['total_ms']) for r in data]
    fps = [float(r['fps']) for r in data]

    latency_results[model] = {
        'frames': len(data),
        'preprocess_ms': round(sum(preprocess) / len(preprocess), 2),
        'inference_ms': round(sum(inference) / len(inference), 2),
        'postprocess_ms': round(sum(postprocess) / len(postprocess), 2),
        'total_ms': round(sum(total) / len(total), 2),
        'fps': round(sum(fps) / len(fps), 2),
    }

# ============================================================
# 2. RESOURCE (Average & Max)
# ============================================================
resource_avg = {}
resource_max = {}

for model in MODELS:
    csv_file = os.path.join(DATA_DIR, f"benchmark_resource_best_{model}_fp16_jetson_nano.csv")
    if not os.path.exists(csv_file):
        continue

    rows = read_csv(csv_file)
    if not rows:
        continue

    cpu = [float(r['cpu_percent']) for r in rows]
    gpu = [float(r['gpu_percent']) for r in rows]
    ram_used = [float(r['ram_used_mb']) for r in rows]
    ram_total = [float(r['ram_total_mb']) for r in rows]
    temp = [float(r['temperature_c']) for r in rows]

    n = len(rows)
    resource_avg[model] = {
        'samples': n,
        'cpu_percent': round(sum(cpu) / n, 2),
        'gpu_percent': round(sum(gpu) / n, 2),
        'ram_used_mb': round(sum(ram_used) / n, 2),
        'ram_total_mb': round(sum(ram_total) / n, 2),
        'temperature_c': round(sum(temp) / n, 2),
    }
    resource_max[model] = {
        'samples': n,
        'cpu_percent': round(max(cpu), 2),
        'gpu_percent': round(max(gpu), 2),
        'ram_used_mb': round(max(ram_used), 2),
        'ram_total_mb': round(max(ram_total), 2),
        'temperature_c': round(max(temp), 2),
    }

# ============================================================
# 3. POWER (Average & Max)
# ============================================================
power_avg = {}
power_max = {}

for model in MODELS:
    csv_file = os.path.join(DATA_DIR, f"power_log_{model}_jetson_nano.csv")
    if not os.path.exists(csv_file):
        continue

    rows = read_csv(csv_file)
    if not rows:
        continue

    power_total = [float(r['power_total_mw']) for r in rows]
    power_gpu = [float(r['power_gpu_mw']) for r in rows]
    power_cpu = [float(r['power_cpu_mw']) for r in rows]

    n = len(rows)
    power_avg[model] = {
        'samples': n,
        'power_total_mw': round(sum(power_total) / n, 2),
        'power_gpu_mw': round(sum(power_gpu) / n, 2),
        'power_cpu_mw': round(sum(power_cpu) / n, 2),
    }
    power_max[model] = {
        'samples': n,
        'power_total_mw': round(max(power_total), 2),
        'power_gpu_mw': round(max(power_gpu), 2),
        'power_cpu_mw': round(max(power_cpu), 2),
    }

# ============================================================
# SAVE CSV + JSON
# ============================================================

# Latency CSV
with open(os.path.join(OUTPUT_DIR, "average_latency_fps_jetson.csv"), 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['model', 'frames', 'preprocess_ms', 'inference_ms', 'postprocess_ms', 'total_ms', 'fps'])
    for m in MODELS:
        if m in latency_results:
            d = latency_results[m]
            w.writerow([m, d['frames'], d['preprocess_ms'], d['inference_ms'],
                       d['postprocess_ms'], d['total_ms'], d['fps']])

# Resource AVG CSV
with open(os.path.join(OUTPUT_DIR, "average_resource_jetson.csv"), 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['model', 'samples', 'cpu_percent', 'gpu_percent', 'ram_used_mb', 'ram_total_mb', 'temperature_c'])
    for m in MODELS:
        if m in resource_avg:
            d = resource_avg[m]
            w.writerow([m, d['samples'], d['cpu_percent'], d['gpu_percent'],
                       d['ram_used_mb'], d['ram_total_mb'], d['temperature_c']])

# Resource MAX CSV
with open(os.path.join(OUTPUT_DIR, "max_resource_jetson.csv"), 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['model', 'samples', 'cpu_percent', 'gpu_percent', 'ram_used_mb', 'ram_total_mb', 'temperature_c'])
    for m in MODELS:
        if m in resource_max:
            d = resource_max[m]
            w.writerow([m, d['samples'], d['cpu_percent'], d['gpu_percent'],
                       d['ram_used_mb'], d['ram_total_mb'], d['temperature_c']])

# Power AVG CSV
with open(os.path.join(OUTPUT_DIR, "average_power_jetson.csv"), 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['model', 'samples', 'power_total_mw', 'power_gpu_mw', 'power_cpu_mw'])
    for m in MODELS:
        if m in power_avg:
            d = power_avg[m]
            w.writerow([m, d['samples'], d['power_total_mw'], d['power_gpu_mw'], d['power_cpu_mw']])

# Power MAX CSV
with open(os.path.join(OUTPUT_DIR, "max_power_jetson.csv"), 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['model', 'samples', 'power_total_mw', 'power_gpu_mw', 'power_cpu_mw'])
    for m in MODELS:
        if m in power_max:
            d = power_max[m]
            w.writerow([m, d['samples'], d['power_total_mw'], d['power_gpu_mw'], d['power_cpu_mw']])

# Combined JSON
combined = {
    'average_latency_fps': latency_results,
    'average_resource': resource_avg,
    'max_resource': resource_max,
    'average_power': power_avg,
    'max_power': power_max,
}
with open(os.path.join(OUTPUT_DIR, "benchmark_summary_jetson.json"), 'w', encoding='utf-8') as f:
    json.dump(combined, f, indent=2, ensure_ascii=False)

# ============================================================
# PRINT TABLES
# ============================================================
print("=" * 90)
print("  AVERAGE BENCHMARK LATENCY & FPS - JETSON NANO (excl. warmup)")
print("=" * 90)
print(f"  {'Model':<12s} {'Frames':>7s} {'Preproc(ms)':>12s} {'Infer(ms)':>11s} {'Postproc(ms)':>13s} {'Total(ms)':>11s} {'FPS':>8s}")
print(f"  {'-'*12} {'-'*7} {'-'*12} {'-'*11} {'-'*13} {'-'*11} {'-'*8}")
for m in MODELS:
    if m in latency_results:
        d = latency_results[m]
        print(f"  {m:<12s} {d['frames']:>7d} {d['preprocess_ms']:>12.2f} {d['inference_ms']:>11.2f} "
              f"{d['postprocess_ms']:>13.2f} {d['total_ms']:>11.2f} {d['fps']:>8.2f}")

print(f"\n{'='*90}")
print("  AVERAGE BENCHMARK RESOURCE - JETSON NANO")
print("=" * 90)
print(f"  {'Model':<12s} {'Samples':>8s} {'CPU(%)':>8s} {'GPU(%)':>8s} {'RAM Used(MB)':>13s} {'RAM Total(MB)':>14s} {'Temp(C)':>9s}")
print(f"  {'-'*12} {'-'*8} {'-'*8} {'-'*8} {'-'*13} {'-'*14} {'-'*9}")
for m in MODELS:
    if m in resource_avg:
        d = resource_avg[m]
        print(f"  {m:<12s} {d['samples']:>8d} {d['cpu_percent']:>8.2f} {d['gpu_percent']:>8.2f} "
              f"{d['ram_used_mb']:>13.2f} {d['ram_total_mb']:>14.2f} {d['temperature_c']:>9.2f}")

print(f"\n{'='*90}")
print("  MAX BENCHMARK RESOURCE - JETSON NANO")
print("=" * 90)
print(f"  {'Model':<12s} {'Samples':>8s} {'CPU(%)':>8s} {'GPU(%)':>8s} {'RAM Used(MB)':>13s} {'RAM Total(MB)':>14s} {'Temp(C)':>9s}")
print(f"  {'-'*12} {'-'*8} {'-'*8} {'-'*8} {'-'*13} {'-'*14} {'-'*9}")
for m in MODELS:
    if m in resource_max:
        d = resource_max[m]
        print(f"  {m:<12s} {d['samples']:>8d} {d['cpu_percent']:>8.2f} {d['gpu_percent']:>8.2f} "
              f"{d['ram_used_mb']:>13.2f} {d['ram_total_mb']:>14.2f} {d['temperature_c']:>9.2f}")

print(f"\n{'='*90}")
print("  AVERAGE POWER - JETSON NANO")
print("=" * 90)
print(f"  {'Model':<12s} {'Samples':>8s} {'Total(mW)':>11s} {'GPU(mW)':>9s} {'CPU(mW)':>9s}")
print(f"  {'-'*12} {'-'*8} {'-'*11} {'-'*9} {'-'*9}")
for m in MODELS:
    if m in power_avg:
        d = power_avg[m]
        print(f"  {m:<12s} {d['samples']:>8d} {d['power_total_mw']:>11.2f} {d['power_gpu_mw']:>9.2f} {d['power_cpu_mw']:>9.2f}")

print(f"\n{'='*90}")
print("  MAX POWER - JETSON NANO")
print("=" * 90)
print(f"  {'Model':<12s} {'Samples':>8s} {'Total(mW)':>11s} {'GPU(mW)':>9s} {'CPU(mW)':>9s}")
print(f"  {'-'*12} {'-'*8} {'-'*11} {'-'*9} {'-'*9}")
for m in MODELS:
    if m in power_max:
        d = power_max[m]
        print(f"  {m:<12s} {d['samples']:>8d} {d['power_total_mw']:>11.2f} {d['power_gpu_mw']:>9.2f} {d['power_cpu_mw']:>9.2f}")

# ============================================================
# VALIDASI vs Excel Manual
# ============================================================
print(f"\n\n{'='*90}")
print("  VALIDASI vs HASIL MANUAL EXCEL (FIXED-baru)")
print("=" * 90)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("  [WARN] openpyxl not installed, skipping Excel validation")
    print("         Install: pip install openpyxl")

if HAS_OPENPYXL:
    for model in MODELS:
        # --- Latency validation ---
        xl_lat = os.path.join(EXCEL_DIR, f"benchmark_fps_latency_best_{model}_fp16_jetson_nano.xlsx")
        if os.path.exists(xl_lat) and model in latency_results:
            wb = openpyxl.load_workbook(xl_lat, data_only=True)
            ws = wb.active

            # Find the summary/average row - look for cells with "AVERAGE" or last rows
            # Try reading all data to find average calculations
            headers = [cell.value for cell in ws[1]]
            
            # Read all non-warmup data from Excel
            xl_preproc = []
            xl_infer = []
            xl_postproc = []
            xl_total = []
            xl_fps = []
            
            warmup_col = None
            for i, h in enumerate(headers):
                if h and 'warmup' in str(h).lower():
                    warmup_col = i
                    break
            
            pre_col = headers.index('preprocess_ms') if 'preprocess_ms' in headers else None
            inf_col = headers.index('inference_ms') if 'inference_ms' in headers else None
            post_col = headers.index('postprocess_ms') if 'postprocess_ms' in headers else None
            tot_col = headers.index('total_ms') if 'total_ms' in headers else None
            fps_col = headers.index('fps') if 'fps' in headers else None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip warmup
                if warmup_col is not None and row[warmup_col] in ('true', True, 'TRUE'):
                    continue
                # Skip non-data rows
                if row[0] is None:
                    continue
                try:
                    if pre_col is not None: xl_preproc.append(float(row[pre_col]))
                    if inf_col is not None: xl_infer.append(float(row[inf_col]))
                    if post_col is not None: xl_postproc.append(float(row[post_col]))
                    if tot_col is not None: xl_total.append(float(row[tot_col]))
                    if fps_col is not None: xl_fps.append(float(row[fps_col]))
                except (TypeError, ValueError):
                    continue
            
            wb.close()
            
            if xl_preproc:
                xl_avg = {
                    'preprocess_ms': round(sum(xl_preproc)/len(xl_preproc), 2),
                    'inference_ms': round(sum(xl_infer)/len(xl_infer), 2),
                    'postprocess_ms': round(sum(xl_postproc)/len(xl_postproc), 2),
                    'total_ms': round(sum(xl_total)/len(xl_total), 2),
                    'fps': round(sum(xl_fps)/len(xl_fps), 2),
                }
                my = latency_results[model]
                
                print(f"\n  --- {model.upper()} LATENCY ---")
                print(f"  {'Metric':<16s} {'Script':>12s} {'Excel Data':>12s} {'Diff':>10s} {'Status':>8s}")
                print(f"  {'-'*16} {'-'*12} {'-'*12} {'-'*10} {'-'*8}")
                
                for key in ['preprocess_ms', 'inference_ms', 'postprocess_ms', 'total_ms', 'fps']:
                    script_val = my[key]
                    excel_val = xl_avg[key]
                    diff = abs(script_val - excel_val)
                    # Tolerance: 0.5% or 0.1 absolute
                    max_val = max(abs(script_val), abs(excel_val), 0.001)
                    pct_diff = diff / max_val * 100
                    status = "✓ OK" if pct_diff < 0.5 else f"⚠ {pct_diff:.1f}%"
                    print(f"  {key:<16s} {script_val:>12.2f} {excel_val:>12.2f} {diff:>10.2f} {status:>8s}")
        
        # --- Resource validation ---
        xl_res = os.path.join(EXCEL_DIR, f"benchmark_resource_best_{model}_fp16_jetson_nano.xlsx")
        if os.path.exists(xl_res) and model in resource_avg:
            wb = openpyxl.load_workbook(xl_res, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            xl_cpu = []
            xl_gpu = []
            xl_ram = []
            xl_temp = []
            
            cpu_col = headers.index('cpu_percent') if 'cpu_percent' in headers else None
            gpu_col = headers.index('gpu_percent') if 'gpu_percent' in headers else None
            ram_col = headers.index('ram_used_mb') if 'ram_used_mb' in headers else None
            temp_col = headers.index('temperature_c') if 'temperature_c' in headers else None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                try:
                    if cpu_col is not None: xl_cpu.append(float(row[cpu_col]))
                    if gpu_col is not None: xl_gpu.append(float(row[gpu_col]))
                    if ram_col is not None: xl_ram.append(float(row[ram_col]))
                    if temp_col is not None: xl_temp.append(float(row[temp_col]))
                except (TypeError, ValueError):
                    continue
            
            wb.close()
            
            if xl_cpu:
                n = len(xl_cpu)
                xl_r_avg = {
                    'cpu_percent': round(sum(xl_cpu)/n, 2),
                    'gpu_percent': round(sum(xl_gpu)/n, 2),
                    'ram_used_mb': round(sum(xl_ram)/n, 2),
                    'temperature_c': round(sum(xl_temp)/n, 2),
                }
                xl_r_max = {
                    'cpu_percent': round(max(xl_cpu), 2),
                    'gpu_percent': round(max(xl_gpu), 2),
                    'ram_used_mb': round(max(xl_ram), 2),
                    'temperature_c': round(max(xl_temp), 2),
                }
                my_avg = resource_avg[model]
                my_max = resource_max[model]
                
                print(f"\n  --- {model.upper()} RESOURCE (AVG) ---")
                print(f"  {'Metric':<16s} {'Script':>12s} {'Excel Data':>12s} {'Diff':>10s} {'Status':>8s}")
                print(f"  {'-'*16} {'-'*12} {'-'*12} {'-'*10} {'-'*8}")
                for key in ['cpu_percent', 'gpu_percent', 'ram_used_mb', 'temperature_c']:
                    sv = my_avg[key]
                    ev = xl_r_avg[key]
                    diff = abs(sv - ev)
                    max_val = max(abs(sv), abs(ev), 0.001)
                    pct = diff / max_val * 100
                    status = "✓ OK" if pct < 0.5 else f"⚠ {pct:.1f}%"
                    print(f"  {key:<16s} {sv:>12.2f} {ev:>12.2f} {diff:>10.2f} {status:>8s}")

                print(f"\n  --- {model.upper()} RESOURCE (MAX) ---")
                print(f"  {'Metric':<16s} {'Script':>12s} {'Excel Data':>12s} {'Diff':>10s} {'Status':>8s}")
                print(f"  {'-'*16} {'-'*12} {'-'*12} {'-'*10} {'-'*8}")
                for key in ['cpu_percent', 'gpu_percent', 'ram_used_mb', 'temperature_c']:
                    sv = my_max[key]
                    ev = xl_r_max[key]
                    diff = abs(sv - ev)
                    max_val = max(abs(sv), abs(ev), 0.001)
                    pct = diff / max_val * 100
                    status = "✓ OK" if pct < 0.5 else f"⚠ {pct:.1f}%"
                    print(f"  {key:<16s} {sv:>12.2f} {ev:>12.2f} {diff:>10.2f} {status:>8s}")

print(f"\n{'='*90}")
print(f"  Output saved to: {OUTPUT_DIR}")
print(f"{'='*90}")
