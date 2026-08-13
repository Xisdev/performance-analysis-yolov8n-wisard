"""
==============================================================================
  01. Generate Tabel Hasil Pelatihan Model - Performa Baseline
==============================================================================
  Script ini membaca results.csv dari setiap varian YOLOv8 (n/s/m/l/x)
  dan menghasilkan tabel performa baseline berisi:
    - Presisi (%)
    - Recall (%)
    - F1-Score (%)
    - mAP50 (%)

  Data diambil dari epoch dengan mAP50 TERTINGGI (best epoch),
  yang merupakan epoch dimana best.pt disimpan.

  Output:
    1. Tabel di terminal (ASCII)
    2. tabel_baseline_performa.csv
    3. tabel_baseline_performa.json
    4. (opsional) Update sheet di tabel_data.xlsx jika file ada

  Cara pakai:
    python "01. generate_tabel_baseline.py"
    python "01. generate_tabel_baseline.py" --train-dir D:\path\to\train_result
==============================================================================
"""

import os
import sys
import csv
import json
import argparse
from collections import OrderedDict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Default path ke folder train_result
DEFAULT_TRAIN_DIR = os.path.join(SCRIPT_DIR, 'train_result')

# Mapping varian
VARIANTS = OrderedDict([
    ('yolov8n', {'folder': 'yolov8n_wisard_ir', 'label': 'YOLOv8n (Nano)'}),
    ('yolov8s', {'folder': 'yolov8s_wisard_ir', 'label': 'YOLOv8s (Small)'}),
    ('yolov8m', {'folder': 'yolov8m_wisard_ir', 'label': 'YOLOv8m (Medium)'}),
    ('yolov8l', {'folder': 'yolov8l_wisard_ir', 'label': 'YOLOv8l (Large)'}),
    ('yolov8x', {'folder': 'yolov8x_wisard_ir', 'label': 'YOLOv8x (X-Large)'}),
])

# Kolom di results.csv Ultralytics
COL_EPOCH = 'epoch'
COL_PRECISION = 'metrics/precision(B)'
COL_RECALL = 'metrics/recall(B)'
COL_MAP50 = 'metrics/mAP50(B)'
COL_MAP5095 = 'metrics/mAP50-95(B)'


def parse_results_csv(csv_path):
    """Baca results.csv dan return list of dict per epoch."""
    rows = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Strip whitespace dari keys (Ultralytics kadang ada spasi)
            cleaned = {}
            for k, v in row.items():
                k_clean = k.strip()
                try:
                    cleaned[k_clean] = float(v.strip())
                except (ValueError, AttributeError):
                    cleaned[k_clean] = v.strip() if isinstance(v, str) else v
            rows.append(cleaned)
    return rows


def find_best_epoch(rows):
    """Cari epoch dengan mAP50 tertinggi."""
    best_row = None
    best_map50 = -1

    for row in rows:
        map50 = row.get(COL_MAP50, 0)
        if isinstance(map50, (int, float)) and map50 > best_map50:
            best_map50 = map50
            best_row = row

    return best_row


def compute_f1(precision, recall):
    """Hitung F1-Score dari Precision dan Recall."""
    if precision + recall == 0:
        return 0.0
    return 2 * (precision * recall) / (precision + recall)


def main():
    parser = argparse.ArgumentParser(
        description="Generate tabel baseline performa dari hasil training YOLOv8"
    )
    parser.add_argument(
        '--train-dir', type=str, default=DEFAULT_TRAIN_DIR,
        help=f"Path ke folder hasil training (default: {DEFAULT_TRAIN_DIR})"
    )
    args = parser.parse_args()

    train_dir = args.train_dir

    print("=" * 75)
    print("  TABEL HASIL PELATIHAN MODEL - PERFORMA BASELINE")
    print("=" * 75)
    print(f"\n  Sumber data: {train_dir}")

    if not os.path.exists(train_dir):
        print(f"\n  [ERROR] Folder tidak ditemukan: {train_dir}")
        return

    # --- Kumpulkan data dari setiap varian ---
    results = []

    for variant, info in VARIANTS.items():
        csv_path = os.path.join(train_dir, info['folder'], 'results.csv')

        if not os.path.exists(csv_path):
            print(f"\n  [SKIP] {variant}: results.csv tidak ditemukan")
            results.append({
                'variant': variant,
                'label': info['label'],
                'status': 'NOT_FOUND',
            })
            continue

        rows = parse_results_csv(csv_path)
        total_epochs = len(rows)
        best = find_best_epoch(rows)

        if best is None:
            print(f"\n  [ERROR] {variant}: Tidak ada data valid di results.csv")
            results.append({
                'variant': variant,
                'label': info['label'],
                'status': 'NO_DATA',
            })
            continue

        precision = best.get(COL_PRECISION, 0)
        recall = best.get(COL_RECALL, 0)
        map50 = best.get(COL_MAP50, 0)
        map5095 = best.get(COL_MAP5095, 0)
        f1 = compute_f1(precision, recall)
        best_epoch = int(best.get(COL_EPOCH, 0))

        results.append({
            'variant': variant,
            'label': info['label'],
            'status': 'OK',
            'total_epochs': total_epochs,
            'best_epoch': best_epoch,
            'precision': round(precision * 100, 2),
            'recall': round(recall * 100, 2),
            'f1_score': round(f1 * 100, 2),
            'map50': round(map50 * 100, 2),
            'map50_95': round(map5095 * 100, 2),
            'precision_raw': precision,
            'recall_raw': recall,
            'f1_raw': f1,
            'map50_raw': map50,
            'map50_95_raw': map5095,
        })

    # --- Tampilkan tabel di terminal ---
    print(f"\n{'=' * 75}")
    print(f"  TABEL 4.1 - PERFORMA BASELINE (Best Epoch)")
    print(f"{'=' * 75}")
    print()

    # Header
    header = f"  {'Model':<20} {'Epoch':>6} {'Presisi':>10} {'Recall':>10} {'F1-Score':>10} {'mAP50':>10} {'mAP50-95':>10}"
    print(header)
    print(f"  {'-'*20} {'-'*6} {'-'*10} {'-'*10} {'-'*10} {'-'*10} {'-'*10}")

    for r in results:
        if r['status'] != 'OK':
            print(f"  {r['label']:<20} {'N/A':>6} {'N/A':>10} {'N/A':>10} {'N/A':>10} {'N/A':>10} {'N/A':>10}")
            continue

        # Tandai jika early stopping (total < 400)
        epoch_str = f"{r['best_epoch']}/{r['total_epochs']}"

        print(f"  {r['label']:<20} {epoch_str:>6} {r['precision']:>9.2f}% {r['recall']:>9.2f}% {r['f1_score']:>9.2f}% {r['map50']:>9.2f}% {r['map50_95']:>9.2f}%")

    print()

    # --- Info tambahan ---
    print(f"  Catatan:")
    for r in results:
        if r['status'] == 'OK':
            es = "(Early Stopping)" if r['total_epochs'] < 400 else "(Full 400 epoch)"
            print(f"    {r['variant']}: best di epoch {r['best_epoch']} dari {r['total_epochs']} {es}")

    # --- Simpan CSV ---
    csv_output = os.path.join(SCRIPT_DIR, 'tabel_baseline_performa.csv')
    with open(csv_output, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Model', 'Varian', 'Total Epoch', 'Best Epoch',
            'Presisi (%)', 'Recall (%)', 'F1-Score (%)', 'mAP50 (%)', 'mAP50-95 (%)'
        ])
        for r in results:
            if r['status'] != 'OK':
                writer.writerow([r['label'], r['variant'], 'N/A', 'N/A',
                                 'N/A', 'N/A', 'N/A', 'N/A', 'N/A'])
            else:
                writer.writerow([
                    r['label'], r['variant'], r['total_epochs'], r['best_epoch'],
                    r['precision'], r['recall'], r['f1_score'], r['map50'], r['map50_95']
                ])

    print(f"\n  CSV disimpan: {csv_output}")

    # --- Simpan JSON ---
    json_output = os.path.join(SCRIPT_DIR, 'tabel_baseline_performa.json')
    json_data = {
        'title': 'Tabel 4.1 - Hasil Pelatihan Model: Performa Baseline',
        'description': 'Metrik performa model YOLOv8 pada dataset WiSARD IR (best epoch)',
        'dataset': 'WiSARD IR (31554 train / 16302 val / 5950 test / 1 class: human)',
        'optimizer': 'AdamW (lr=0.001, wd=0.01)',
        'models': []
    }

    for r in results:
        if r['status'] == 'OK':
            json_data['models'].append({
                'variant': r['variant'],
                'label': r['label'],
                'total_epochs': r['total_epochs'],
                'best_epoch': r['best_epoch'],
                'early_stopping': r['total_epochs'] < 400,
                'metrics': {
                    'precision_pct': r['precision'],
                    'recall_pct': r['recall'],
                    'f1_score_pct': r['f1_score'],
                    'mAP50_pct': r['map50'],
                    'mAP50_95_pct': r['map50_95'],
                }
            })

    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)

    print(f"  JSON disimpan: {json_output}")

    # --- Update Excel jika ada ---
    xlsx_path = os.path.join(SCRIPT_DIR, 'tabel_data.xlsx')
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)

            # Cari atau buat sheet "Tabel 4.1"
            sheet_name = 'Tabel 4.1 - Baseline'
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name, 0)

            # Header
            headers = ['Model', 'Total Epoch', 'Best Epoch',
                        'Presisi (%)', 'Recall (%)', 'F1-Score (%)',
                        'mAP50 (%)', 'mAP50-95 (%)']
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)

            # Data
            row_idx = 2
            for r in results:
                if r['status'] != 'OK':
                    ws.cell(row=row_idx, column=1, value=r['label'])
                    ws.cell(row=row_idx, column=2, value='N/A')
                else:
                    ws.cell(row=row_idx, column=1, value=r['label'])
                    ws.cell(row=row_idx, column=2, value=r['total_epochs'])
                    ws.cell(row=row_idx, column=3, value=r['best_epoch'])
                    ws.cell(row=row_idx, column=4, value=r['precision'])
                    ws.cell(row=row_idx, column=5, value=r['recall'])
                    ws.cell(row=row_idx, column=6, value=r['f1_score'])
                    ws.cell(row=row_idx, column=7, value=r['map50'])
                    ws.cell(row=row_idx, column=8, value=r['map50_95'])
                row_idx += 1

            wb.save(xlsx_path)
            print(f"  Excel diupdate: {xlsx_path} (sheet: {sheet_name})")
        except ImportError:
            print(f"  [WARN] openpyxl tidak terinstall, Excel tidak diupdate")
        except Exception as e:
            print(f"  [ERROR] Gagal update Excel: {e}")

    print(f"\n{'=' * 75}")


if __name__ == "__main__":
    main()
