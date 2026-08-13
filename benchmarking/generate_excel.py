"""
==============================================================================
  Update tabel_data.xlsx dari Hasil Benchmark
==============================================================================
  Script ini membaca file JSON/CSV hasil benchmark dari folder jetson_nano/
  dan raspberry_pi_4/, lalu mengisi data ke sheet tabel_data.xlsx.

  Cara pakai:
      python generate_excel.py --update

  Input:
      jetson_nano/
      ├── benchmark_summary_{model}_jetson_nano.json   --> FPS & Latency
      ├── benchmark_resource_{model}_jetson_nano.csv    --> Resource usage
      └── eval_accuracy_{model}_jetson_nano.json        --> Akurasi

      raspberry_pi_4/
      ├── benchmark_summary_{model}_raspi4.json         --> FPS & Latency
      ├── benchmark_resource_{model}_raspi4.csv          --> Resource usage
      └── eval_accuracy_{model}_raspi4.json              --> Akurasi

  Output:
      D:\RISET\drone-wisard\penulisan\tabel_data.xlsx
        Sheet "fps&latency ke2perangkat"        (Tabel 4.2)
        Sheet "eval akurasi setelah convert"     (Tabel 4.3)
        Sheet "Penggunaan resource perangkat"    (Tabel 4.4)
==============================================================================
"""

import os
import sys
import csv
import json
import glob
import argparse
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[ERROR] openpyxl belum terinstall!")
    print("        Jalankan: pip install openpyxl")
    sys.exit(1)


# ============================================================================
#  KONFIGURASI
# ============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Path ke file Excel yang sudah ada
EXCEL_PATH = os.path.join(os.path.dirname(SCRIPT_DIR), "penulisan", "tabel_data.xlsx")

# Folder hasil benchmark per device
JETSON_DIR = os.path.join(SCRIPT_DIR, "jetson_nano")
RASPI_DIR = os.path.join(SCRIPT_DIR, "raspberry_pi_4")

# Nama sheet di tabel_data.xlsx (HARUS SAMA PERSIS)
SHEET_FPS_LATENCY = "fps&latency ke2perangkat"
SHEET_EVAL_AKURASI = "eval akurasi setelah convert"
SHEET_RESOURCE = "Penggunaan resource perangkat"

# Mapping varian model ke urutan baris
# Index = urutan model (0=n, 1=s, 2=m, 3=l, 4=x)
MODEL_VARIANTS = ['yolov8n', 'yolov8s', 'yolov8m', 'yolov8l', 'yolov8x']
MODEL_LABELS = ['YOLOv8n', 'YOLOv8s', 'YOLOv8m', 'YOLOv8l', 'YOLOv8x']

# ============================================================================
#  Sheet "fps&latency ke2perangkat" - Layout:
#
#  Row 1: Header baris 1 (Perangkat, Akselerator, Varian, Latensi, FPS, Status)
#  Row 2: Sub-header (Preprocess, Inferensi, Postprocess, Total)
#  Row 3-7: Jetson Nano (n, s, m, l, x)
#  Row 8-12: Raspberry Pi 4 (n, s, m, l, x)
#
#  Kolom:
#    C1=Perangkat, C2=Akselerator, C3=Varian
#    C4=Preprocess(ms), C5=Inferensi(ms), C6=Postprocess(ms), C7=Total(ms)
#    C8=FPS, C9=Status
# ============================================================================

# ============================================================================
#  Sheet "eval akurasi setelah convert" - Layout:
#
#  Row 1: Header (Varian Model, Metrik, Baseline PC, Jetson Nano, RPi4, Penurunan)
#  Row 2-5:   YOLOv8n (Precision, Recall, F1, mAP50)
#  Row 6-9:   YOLOv8s
#  Row 10-13: YOLOv8m
#  Row 14-17: YOLOv8l
#  Row 18-21: YOLOv8x
#
#  Kolom:
#    C1=Varian Model, C2=Metrik
#    C3=Baseline PC (.pt)
#    C4=Jetson Nano (.engine FP16)
#    C5=Raspberry Pi 4 (.tflite FP32)
#    C6=Penurunan
# ============================================================================

# ============================================================================
#  Sheet "Penggunaan resource perangkat" - Layout:
#
#  Row 1: Header
#  Row 2-6:  Jetson Nano (n, s, m, l, x)
#  Row 7-11: Raspberry Pi 4 (n, s, m, l, x)
#
#  Kolom:
#    C1=Perangkat, C2=Varian model
#    C3=CPU Avg %, C4=CPU Max %
#    C5=GPU Avg %, C6=GPU Max %
#    C7=RAM Avg (MB), C8=RAM Max (MB), C9=RAM Total (MB)
#    C10=Suhu Avg (°C), C11=Suhu MAX (°C)
# ============================================================================


def find_variant_index(model_name):
    """Cari index varian model (0-4) dari nama file model."""
    name_lower = model_name.lower()
    for idx, variant in enumerate(MODEL_VARIANTS):
        if variant in name_lower:
            return idx
    # Juga cek singkatan
    for idx, suffix in enumerate(['8n', '8s', '8m', '8l', '8x']):
        if suffix in name_lower:
            return idx
    return -1


def load_benchmark_summaries(directory, device_key):
    """Load semua benchmark_summary_*.json dari folder."""
    results = {}
    pattern = os.path.join(directory, f"benchmark_summary_*_{device_key}.json")
    files = sorted(glob.glob(pattern))

    for filepath in files:
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            model_name = data.get('model_name', Path(filepath).stem)
            idx = find_variant_index(model_name)
            if idx >= 0:
                results[idx] = data
                print(f"  [OK] Loaded: {os.path.basename(filepath)} -> {MODEL_LABELS[idx]}")
            else:
                print(f"  [SKIP] Tidak bisa mapping varian: {os.path.basename(filepath)}")
        except Exception as e:
            print(f"  [ERROR] {filepath}: {e}")

    return results


def load_eval_accuracy(directory, device_key):
    """Load semua eval_accuracy_*.json dari folder."""
    results = {}
    pattern = os.path.join(directory, f"eval_accuracy_*_{device_key}.json")
    files = sorted(glob.glob(pattern))

    for filepath in files:
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            model_name = data.get('model_name', Path(filepath).stem)
            idx = find_variant_index(model_name)
            if idx >= 0:
                results[idx] = data
                print(f"  [OK] Loaded: {os.path.basename(filepath)} -> {MODEL_LABELS[idx]}")
            else:
                print(f"  [SKIP] Tidak bisa mapping varian: {os.path.basename(filepath)}")
        except Exception as e:
            print(f"  [ERROR] {filepath}: {e}")

    return results


def load_resource_csv(directory, device_key):
    """Load dan agregasi benchmark_resource_*.csv dari folder."""
    results = {}
    pattern = os.path.join(directory, f"benchmark_resource_*_{device_key}.csv")
    files = sorted(glob.glob(pattern))

    for filepath in files:
        try:
            rows = []
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        rows.append({
                            'cpu_percent': float(row['cpu_percent']),
                            'gpu_percent': float(row['gpu_percent']),
                            'ram_used_mb': float(row['ram_used_mb']),
                            'ram_total_mb': float(row['ram_total_mb']),
                            'temperature_c': float(row['temperature_c']),
                        })
                    except (ValueError, KeyError):
                        continue

            if not rows:
                continue

            basename = Path(filepath).stem
            idx = find_variant_index(basename)
            if idx >= 0:
                # Agregasi: hitung avg dan max
                import statistics
                cpus = [r['cpu_percent'] for r in rows]
                gpus = [r['gpu_percent'] for r in rows]
                rams = [r['ram_used_mb'] for r in rows]
                temps = [r['temperature_c'] for r in rows]

                results[idx] = {
                    'cpu_avg': round(statistics.mean(cpus), 1),
                    'cpu_max': round(max(cpus), 1),
                    'gpu_avg': round(statistics.mean(gpus), 1),
                    'gpu_max': round(max(gpus), 1),
                    'ram_avg': round(statistics.mean(rams), 0),
                    'ram_max': round(max(rams), 0),
                    'ram_total': round(rows[0]['ram_total_mb'], 0),
                    'temp_avg': round(statistics.mean(temps), 1),
                    'temp_max': round(max(temps), 1),
                }
                print(f"  [OK] Loaded: {os.path.basename(filepath)} -> {MODEL_LABELS[idx]} ({len(rows)} samples)")
        except Exception as e:
            print(f"  [ERROR] {filepath}: {e}")

    return results


def update_fps_latency(wb, jetson_summaries, raspi_summaries):
    """Update Sheet 'fps&latency ke2perangkat'."""
    if SHEET_FPS_LATENCY not in wb.sheetnames:
        print(f"\n  [SKIP] Sheet '{SHEET_FPS_LATENCY}' tidak ditemukan!")
        return 0

    ws = wb[SHEET_FPS_LATENCY]
    count = 0

    # Jetson Nano: Row 3-7 (index 0-4)
    for idx, data in jetson_summaries.items():
        row = 3 + idx  # Row 3=n, 4=s, 5=m, 6=l, 7=x
        if data.get('status', 'OK') == 'OK' and 'latency' in data:
            ws.cell(row=row, column=4, value=round(data['latency']['preprocess_ms']['mean'], 2))
            ws.cell(row=row, column=5, value=round(data['latency']['inference_ms']['mean'], 2))
            ws.cell(row=row, column=6, value=round(data['latency']['postprocess_ms']['mean'], 2))
            ws.cell(row=row, column=7, value=round(data['latency']['total_ms']['mean'], 2))
            ws.cell(row=row, column=8, value=round(data['fps']['mean'], 2))
            ws.cell(row=row, column=9, value='OK')
            print(f"    [v] Jetson {MODEL_LABELS[idx]}: FPS={data['fps']['mean']:.2f}, Latency={data['latency']['total_ms']['mean']:.2f}ms")
            count += 1
        else:
            status = data.get('status', 'ERROR')
            ws.cell(row=row, column=9, value=status)
            print(f"    [v] Jetson {MODEL_LABELS[idx]}: {status}")
            count += 1

    # Raspberry Pi 4: Row 8-12 (index 0-4)
    for idx, data in raspi_summaries.items():
        row = 8 + idx  # Row 8=n, 9=s, 10=m, 11=l, 12=x
        if data.get('status', 'OK') == 'OK' and 'latency' in data:
            ws.cell(row=row, column=4, value=round(data['latency']['preprocess_ms']['mean'], 2))
            ws.cell(row=row, column=5, value=round(data['latency']['inference_ms']['mean'], 2))
            ws.cell(row=row, column=6, value=round(data['latency']['postprocess_ms']['mean'], 2))
            ws.cell(row=row, column=7, value=round(data['latency']['total_ms']['mean'], 2))
            ws.cell(row=row, column=8, value=round(data['fps']['mean'], 2))
            ws.cell(row=row, column=9, value='OK')
            print(f"    [v] RPi4 {MODEL_LABELS[idx]}: FPS={data['fps']['mean']:.2f}, Latency={data['latency']['total_ms']['mean']:.2f}ms")
            count += 1
        else:
            status = data.get('status', 'ERROR')
            ws.cell(row=row, column=9, value=status)
            print(f"    [v] RPi4 {MODEL_LABELS[idx]}: {status}")
            count += 1

    return count


def update_eval_akurasi(wb, jetson_evals, raspi_evals):
    """Update Sheet 'eval akurasi setelah convert'."""
    if SHEET_EVAL_AKURASI not in wb.sheetnames:
        print(f"\n  [SKIP] Sheet '{SHEET_EVAL_AKURASI}' tidak ditemukan!")
        return 0

    ws = wb[SHEET_EVAL_AKURASI]
    count = 0

    # Layout: setiap model = 4 baris (Precision, Recall, F1, mAP50)
    # YOLOv8n: Row 2-5, YOLOv8s: Row 6-9, YOLOv8m: Row 10-13, dst.
    metric_keys = ['precision', 'recall', 'f1_score', 'mAP50']
    metric_labels = ['Precision', 'Recall', 'F1', 'mAP50']

    for idx in range(5):  # 5 model variants
        base_row = 2 + (idx * 4)  # Row 2, 6, 10, 14, 18

        # Jetson Nano -> Column D (C4)
        if idx in jetson_evals:
            data = jetson_evals[idx]
            if data.get('status') == 'OK' and 'metrics' in data:
                m = data['metrics']
                for mi, key in enumerate(metric_keys):
                    row = base_row + mi
                    val = m.get(key, None)
                    if val is not None:
                        ws.cell(row=row, column=4, value=round(val, 2))
                print(f"    [v] Jetson {MODEL_LABELS[idx]}: Prec={m.get('precision',0):.2f}, mAP50={m.get('mAP50',0):.2f}")
                count += 1

        # Raspberry Pi 4 -> Column E (C5)
        if idx in raspi_evals:
            data = raspi_evals[idx]
            if data.get('status') == 'OK' and 'metrics' in data:
                m = data['metrics']
                for mi, key in enumerate(metric_keys):
                    row = base_row + mi
                    val = m.get(key, None)
                    if val is not None:
                        ws.cell(row=row, column=5, value=round(val, 2))
                print(f"    [v] RPi4 {MODEL_LABELS[idx]}: Prec={m.get('precision',0):.2f}, mAP50={m.get('mAP50',0):.2f}")
                count += 1

        # Hitung penurunan (kolom F / C6) = rata-rata penurunan dari baseline
        # Penurunan = Baseline - rata-rata(Jetson, RPi4)
        for mi, key in enumerate(metric_keys):
            row = base_row + mi
            baseline = ws.cell(row=row, column=3).value
            jetson_val = ws.cell(row=row, column=4).value
            raspi_val = ws.cell(row=row, column=5).value

            if baseline is not None and (jetson_val is not None or raspi_val is not None):
                try:
                    bl = float(baseline)
                    vals = []
                    if jetson_val is not None and isinstance(jetson_val, (int, float)):
                        vals.append(float(jetson_val))
                    if raspi_val is not None and isinstance(raspi_val, (int, float)):
                        vals.append(float(raspi_val))
                    if vals:
                        avg_device = sum(vals) / len(vals)
                        penurunan = bl - avg_device
                        ws.cell(row=row, column=6, value=round(penurunan, 2))
                except (ValueError, TypeError):
                    pass

    return count


def update_resource(wb, jetson_resources, raspi_resources):
    """Update Sheet 'Penggunaan resource perangkat'."""
    if SHEET_RESOURCE not in wb.sheetnames:
        print(f"\n  [SKIP] Sheet '{SHEET_RESOURCE}' tidak ditemukan!")
        return 0

    ws = wb[SHEET_RESOURCE]
    count = 0

    # Jetson Nano: Row 2-6
    for idx, data in jetson_resources.items():
        row = 2 + idx  # Row 2=n, 3=s, 4=m, 5=l, 6=x
        ws.cell(row=row, column=3, value=data['cpu_avg'])
        ws.cell(row=row, column=4, value=data['cpu_max'])
        ws.cell(row=row, column=5, value=data['gpu_avg'])
        ws.cell(row=row, column=6, value=data['gpu_max'])
        ws.cell(row=row, column=7, value=data['ram_avg'])
        ws.cell(row=row, column=8, value=data['ram_max'])
        # C9 (RAM Total) sudah diisi = 4096, skip
        ws.cell(row=row, column=10, value=data['temp_avg'])
        ws.cell(row=row, column=11, value=data['temp_max'])
        print(f"    [v] Jetson {MODEL_LABELS[idx]}: CPU={data['cpu_avg']}%, RAM={data['ram_avg']}MB, Temp={data['temp_avg']}°C")
        count += 1

    # Raspberry Pi 4: Row 7-11
    for idx, data in raspi_resources.items():
        row = 7 + idx  # Row 7=n, 8=s, 9=m, 10=l, 11=x
        ws.cell(row=row, column=3, value=data['cpu_avg'])
        ws.cell(row=row, column=4, value=data['cpu_max'])
        # C5, C6 (GPU) = N/A untuk RPi4, sudah diisi, skip
        ws.cell(row=row, column=7, value=data['ram_avg'])
        ws.cell(row=row, column=8, value=data['ram_max'])
        # C9 (RAM Total) sudah diisi = 4096, skip
        ws.cell(row=row, column=10, value=data['temp_avg'])
        ws.cell(row=row, column=11, value=data['temp_max'])
        print(f"    [v] RPi4 {MODEL_LABELS[idx]}: CPU={data['cpu_avg']}%, RAM={data['ram_avg']}MB, Temp={data['temp_avg']}°C")
        count += 1

    return count


def main():
    parser = argparse.ArgumentParser(
        description="Update tabel_data.xlsx dari hasil benchmark"
    )
    parser.add_argument(
        "--update", action="store_true",
        help="Update data dari file JSON/CSV hasil benchmark"
    )
    parser.add_argument(
        "--excel", type=str, default=EXCEL_PATH,
        help=f"Path ke file tabel_data.xlsx (default: {EXCEL_PATH})"
    )
    parser.add_argument(
        "--jetson-dir", type=str, default=JETSON_DIR,
        help=f"Folder hasil Jetson Nano (default: {JETSON_DIR})"
    )
    parser.add_argument(
        "--raspi-dir", type=str, default=RASPI_DIR,
        help=f"Folder hasil RPi4 (default: {RASPI_DIR})"
    )

    args = parser.parse_args()

    print("=" * 70)
    print("  UPDATE TABEL DATA EXCEL")
    print("  Mengisi data dari hasil benchmark ke tabel_data.xlsx")
    print("=" * 70)

    if not args.update:
        print("\n[INFO] Gunakan --update untuk mengisi data.")
        print("       Contoh: python generate_excel.py --update")
        parser.print_help()
        return

    excel_path = args.excel
    if not os.path.exists(excel_path):
        print(f"\n[ERROR] File Excel tidak ditemukan: {excel_path}")
        print("        Pastikan tabel_data.xlsx sudah ada.")
        sys.exit(1)

    print(f"\n  Excel : {excel_path}")
    print(f"  Jetson: {args.jetson_dir}")
    print(f"  RPi4  : {args.raspi_dir}")

    # Load data dari file JSON/CSV
    print(f"\n{'-' * 50}")
    print(f"  [1/4] Loading benchmark summaries...")
    print(f"{'-' * 50}")

    print(f"\n  --- Jetson Nano ---")
    jetson_summaries = load_benchmark_summaries(args.jetson_dir, "jetson_nano")

    print(f"\n  --- Raspberry Pi 4 ---")
    raspi_summaries = load_benchmark_summaries(args.raspi_dir, "raspi4")

    print(f"\n{'-' * 50}")
    print(f"  [2/4] Loading accuracy evaluations...")
    print(f"{'-' * 50}")

    print(f"\n  --- Jetson Nano ---")
    jetson_evals = load_eval_accuracy(args.jetson_dir, "jetson_nano")

    print(f"\n  --- Raspberry Pi 4 ---")
    raspi_evals = load_eval_accuracy(args.raspi_dir, "raspi4")

    print(f"\n{'-' * 50}")
    print(f"  [3/4] Loading resource CSVs...")
    print(f"{'-' * 50}")

    print(f"\n  --- Jetson Nano ---")
    jetson_resources = load_resource_csv(args.jetson_dir, "jetson_nano")

    print(f"\n  --- Raspberry Pi 4 ---")
    raspi_resources = load_resource_csv(args.raspi_dir, "raspi4")

    # Ringkasan data yang ditemukan
    total_data = (len(jetson_summaries) + len(raspi_summaries) +
                  len(jetson_evals) + len(raspi_evals) +
                  len(jetson_resources) + len(raspi_resources))

    if total_data == 0:
        print(f"\n[WARNING] Tidak ada data benchmark yang ditemukan!")
        print(f"          Pastikan file JSON/CSV sudah ada di:")
        print(f"          - {args.jetson_dir}")
        print(f"          - {args.raspi_dir}")
        print(f"\n          Format file yang diharapkan:")
        print(f"          - benchmark_summary_*_jetson_nano.json")
        print(f"          - benchmark_summary_*_raspi4.json")
        print(f"          - eval_accuracy_*_jetson_nano.json")
        print(f"          - eval_accuracy_*_raspi4.json")
        print(f"          - benchmark_resource_*_jetson_nano.csv")
        print(f"          - benchmark_resource_*_raspi4.csv")
        return

    # Update Excel
    print(f"\n{'-' * 50}")
    print(f"  [4/4] Updating Excel...")
    print(f"{'-' * 50}")

    wb = openpyxl.load_workbook(excel_path)

    print(f"\n  Sheet yang tersedia: {wb.sheetnames}")

    total_updated = 0

    print(f"\n  --- Updating FPS & Latency ---")
    total_updated += update_fps_latency(wb, jetson_summaries, raspi_summaries)

    print(f"\n  --- Updating Evaluasi Akurasi ---")
    total_updated += update_eval_akurasi(wb, jetson_evals, raspi_evals)

    print(f"\n  --- Updating Resource Usage ---")
    total_updated += update_resource(wb, jetson_resources, raspi_resources)

    # Simpan
    wb.save(excel_path)
    print(f"\n{'=' * 70}")
    print(f"  SELESAI!")
    print(f"  Total {total_updated} data point diupdate.")
    print(f"  File disimpan: {excel_path}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
